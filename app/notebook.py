"""
Este módulo contém a classe Notebook, que é uma subclasse da classe ttk.Notebook.
Ela é responsável por configurar e adicionar frames para as abas de Conex o, Check-in e
Check-out ao notebook. A classe Notebook inicializa as frames para as abas de Conex o,
Check-in e Check-out criando instâncias das respectivas classes e adicionando-as ao widget
notebook com rótulos e estados apropriados.

A classe Notebook também fornece métodos para abrir e salvar arquivos Excel. O método
open_file permite que o usuário selecione um arquivo Excel para abrir, enquanto o método
save_file permite que o usuário salve o arquivo Excel atual. O método save_file também atualiza
o status do arquivo Excel atual na barra de status.
"""

from tkinter import ttk
import os
import sys
from openpyxl import load_workbook
from checkin import AppCheckin
from conexao import Conexao
from checkout import AppCheckout
from PIL import ImageTk, Image


class Notebook(ttk.Notebook):
    def __init__(self, parent):
        """
        Inicializa o objeto Notebook com o pai parent.

        Abre as abas de Conexão, Check-in e Check-out.
        """
        super().__init__(parent)
        self.parent = parent
        self.pack(fill='both', expand=True)
        self.carregar_imagens()
        self.abrir_abas(self)
    def abrir_abas(self, notebook):
        """
        Configures and adds the frames for Conexão, Check-in, and Check-out to the notebook.

        This method initializes the frames for the Conexão, Check-in, and Check-out 
        tabs by creating instances of the respective classes and adding them to the 
        notebook widget with appropriate labels and states.
        """

        self.frame_conexao = Conexao(notebook)
        self.frame_checkin = AppCheckin(notebook)
        self.frame_checkout = AppCheckout(notebook)
        # Abrindo Notebook
        notebook.add(self.frame_conexao, text='Conexão')
        notebook.add(self.frame_checkin, state='normal', text='Check-in')
        notebook.add(self.frame_checkout, state='normal', text='Check-out')
    def carregar_imagens(self):        
        """
        Loads and resizes various icons for the application's UI.

        This method determines the base path for the executable or development
        environment and loads images from the 'data' directory. Each image is
        resized to fit specific UI components and is converted to a PhotoImage
        object for use within the Tkinter interface. Icons include agent, calendar,
        WhatsApp, upload, show, refresh, and send, each with designated sizes.
        """

        if getattr(sys, 'frozen', False):
            # O aplicativo está rodando como um executável
            base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        else:
            # O aplicativo está rodando em um ambiente de desenvolvimento
            base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

        print(base_path)

        # Carrega e redimensiona a imagem de conexão (arquivo PNG)
        imagem_agente = Image.open(os.path.join(base_path, 'data', 'agent.png'))
        imagem_agente = imagem_agente.resize((24, 24), Image.Resampling.LANCZOS)
        self.icone_agente = ImageTk.PhotoImage(imagem_agente)

        # Carrega e redimensiona a imagem de upload (arquivo PNG)
        imagem_calendario = Image.open(os.path.join(base_path, 'data', 'calendar.png'))
        imagem_calendario = imagem_calendario.resize((12, 12), Image.Resampling.LANCZOS)
        self.icone_calendario = ImageTk.PhotoImage(imagem_calendario)

        # Carrega e redimensiona a imagem do WhatsApp (arquivo GIF)
        imagem_whats = Image.open(os.path.join(base_path, 'data', 'whatsapp.gif'))
        imagem_whats = imagem_whats.resize((48, 48), Image.Resampling.LANCZOS)
        self.icone_whatsapp = ImageTk.PhotoImage(imagem_whats)

        # Carrega e redimensiona a imagem de upload (arquivo PNG)
        imagem_upload = Image.open(os.path.join(base_path, 'data', 'upload.png'))
        imagem_upload = imagem_upload.resize((24, 24), Image.Resampling.LANCZOS)
        self.icone_upload = ImageTk.PhotoImage(imagem_upload)

        # Carrega e redimensiona a imagem de visualização (arquivo PNG)
        imagem_show = Image.open(os.path.join(base_path, 'data', 'show.png'))
        imagem_show = imagem_show.resize((24, 24), Image.Resampling.LANCZOS)
        self.icone_show = ImageTk.PhotoImage(imagem_show)

        # Carrega e redimensiona a imagem de visualização (arquivo PNG)
        imagem_refresh = Image.open(os.path.join(base_path, 'data', 'refresh.png'))
        imagem_refresh = imagem_refresh.resize((24, 24), Image.Resampling.LANCZOS)
        self.icone_refresh = ImageTk.PhotoImage(imagem_refresh)

        # Carrega e redimensiona a imagem de visualização (arquivo PNG)
        imagem_send = Image.open(os.path.join(base_path, 'data', 'send.png'))
        imagem_send = imagem_send.resize((24, 24), Image.Resampling.LANCZOS)
        self.icone_send = ImageTk.PhotoImage(imagem_send)
    def atualizar_status(self, texto, cor):
        """
        Atualiza o label de conex o na aba de conex o.

        Se o texto for "Conectado!", habilita todas as abas.
        Caso contr rio, desabilita todas as abas exceto a de conex o.

        :param texto: texto a ser exibido no label de conex o
        :param cor: cor do texto a ser exibido no label de conex o
        """
        self.frame_conexao.label_conexao.config(text=f"Conexão: {texto}", foreground=cor)
        if texto == "Conectado!":
            for tab_id in range(1, self.index("end")):
                self.tab(tab_id, state="normal")
        else:
            for tab_id in range(1, self.index("end")):
                self.tab(tab_id, state="disabled")
            self.select(0)
    def atualizar_planilha(self, caminho, mes, telefone, coluna, status):
        """
        Atualiza a planilha com o status de uma determinada pessoa.

        :param caminho: caminho completo da planilha
        :param mes: nome da sheet a ser atualizada
        :param telefone: telefone da pessoa a ser atualizada
        :param coluna: n mero da coluna a ser atualizada
        :param status: status a ser atualizado
        :return: True se o status for atualizado com sucesso, False caso contr rio
        """
        try:
            wb = load_workbook(caminho)

            if mes not in wb.sheetnames:
                print(f"Sheet {mes} não encontrada na planilha.")
                return False

            ws = wb[mes]

            for row in ws.iter_rows(min_row=12, 
                                    max_row=ws.max_row, min_col=2, max_col=ws.max_column):
                valor_telefone = str(row[2].value)
                if telefone in valor_telefone:
                    row[coluna].value = status
                    print(f"Status atualizado para {telefone} na planilha.")

            wb.save(caminho)                
            return True
        except (OSError, IOError) as e:
            print(f"Erro ao atualizar status na planilha: {str(e)}")
            return False
        